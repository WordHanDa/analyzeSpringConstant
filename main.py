from openpyxl import load_workbook

def find_swing_value(data_name):
    change_count = 0
    prev_positive = False
    for value in data_name:
        if value > 0 and not prev_positive:
            change_count += 1
            prev_positive = True
        elif value <= 0:
            prev_positive = False
    return change_count

def find_start_time(acceleration_data_name, time_data_name):
    swing_time = 0
    swing_time_row = 0
    for index, value in enumerate(acceleration_data_name):
        if value > 1.5:
            swing_time = time_data_name[index]
            swing_time_row = index + 1  # Adding 1 to convert index to row number
            break
    return swing_time, swing_time_row

wb = load_workbook('phyphox 2023-06-05 23-16-08.xlsx')
ws = wb.active

def get_swing_time(start_time, last_time):
    return last_time - start_time

# Store absolute acceleration values in Absolute_acceleration_data list
Absolute_acceleration_data = []
time = []  # Create an empty list to store time values
for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, min_col=5, max_col=5):
    for cell in row:
        # Check if cell value is numeric before subtracting 9.8
        if isinstance(cell.value, (int, float)):
            Absolute_acceleration_data.append(cell.value - 9.8)

# Populate the time list
for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        time.append(cell.value)

# Calculate the number of times values change from positive to negative
swing_times, swing_time_row = find_start_time(Absolute_acceleration_data, time)
start_time = swing_times
last_time = time[-1]  # Get the last time value
swing_time = get_swing_time(start_time, last_time)
swing_num = find_swing_value(Absolute_acceleration_data)
T = swing_time/swing_num
M = 0.1

print("Swing times:", swing_num)
print("Swing time:", swing_time)
print("T= ",T)
print("K= ",4*3.14*3.14*M/T/T)
